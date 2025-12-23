import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Mark Ronnel Nieva\Desktop\TK Observation Tracker.xlsx')
sheet = wb.active

print(f'Sheet: {sheet.title}')
print(f'Rows: {sheet.max_row}, Cols: {sheet.max_column}')
print()

print('=== COLUMN HEADERS ===')
headers = [cell.value for cell in sheet[1]]
for i, h in enumerate(headers, 1):
    print(f'  Column {i}: {h}')

print()
print('=== SAMPLE DATA (First 5 data rows) ===')

for row_idx in range(2, min(7, sheet.max_row + 1)):
    print(f'--- Row {row_idx} ---')
    for col_idx, header in enumerate(headers, 1):
        val = sheet.cell(row=row_idx, column=col_idx).value
        if val is not None and str(val).strip():
            print(f'  {header}: {val}')

# Also show unique values for Type column if exists
print()
print('=== UNIQUE VALUES IN TYPE-RELATED COLUMNS ===')
for col_idx, header in enumerate(headers, 1):
    if header and ('type' in str(header).lower() or 'category' in str(header).lower() or 'observation' in str(header).lower()):
        unique_vals = set()
        for row_idx in range(2, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val:
                unique_vals.add(str(val))
        print(f'{header}: {sorted(unique_vals)}')
